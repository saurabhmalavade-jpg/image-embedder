# core.py — Streamlit wrapper around the original CMD script logic
# Strategy: download ALL images concurrently (like CMD), 
#            but return raw bytes (not openpyxl objects) to stay thread-safe
#            then embed in small batches to keep RAM under Streamlit's ~1GB cap

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import requests
from PIL import Image as PILImage
import io
import concurrent.futures
import os
import time
import gc

# ── Exact same constants as your original script ───────────────────────────────
MAX_WORKERS     = min(32, (os.cpu_count() or 1) * 5)
REQUEST_TIMEOUT = 20
MAX_RETRIES     = 3
BACKOFF         = 0.75
EMBED_CHUNK     = 50   # embed into workbook 50-at-a-time to avoid RAM spike

# ── Exact same headers as your original script ─────────────────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ExcelImageFetcher/1.0; +https://pattern.com)",
    "Accept": "image/*,application/octet-stream;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}


# ── Exact same retry helper as your original script ───────────────────────────
def fetch_bytes_with_retries(url, timeout=REQUEST_TIMEOUT, max_retries=MAX_RETRIES, backoff=BACKOFF):
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            resp.raise_for_status()
            return resp.content
        except Exception as e:
            last_exc = e
            if attempt < max_retries:
                time.sleep(backoff * attempt)
    raise last_exc


# ── Exact same worker as your original script ─────────────────────────────────
# Only change: returns img_bytes instead of openpyxl Image (thread-safety)
def download_and_process_image(task):
    url = task['url']
    try:
        # 1) Download with retries + headers
        raw = fetch_bytes_with_retries(url)

        # 2) Fully decode — avoids verify() trap (same as your script)
        pil_img = PILImage.open(io.BytesIO(raw))
        pil_img.load()

        # 3) Mode conversion — handles WEBP/HEIC/CMYK/P etc. (same as your script)
        if pil_img.mode in ("P", "LA"):
            pil_img = pil_img.convert("RGBA")
        elif pil_img.mode not in ("RGB", "RGBA"):
            pil_img = pil_img.convert("RGB")

        # 4) Re-encode to PNG — same as your script
        out_png = io.BytesIO()
        pil_img.save(out_png, format="PNG", optimize=True)
        out_png.seek(0)
        img_bytes = out_png.getvalue()

        # 5) Sizing — exact same as your script
        aspect_ratio = pil_img.height / pil_img.width if pil_img.width else 1.0
        img_w = 150
        img_h = img_w * aspect_ratio

        return {
            'status':     'success',
            'task':       task,
            'img_bytes':  img_bytes,   # bytes instead of Image object
            'img_w':      img_w,
            'img_h':      img_h,
            'row_height': img_h * 0.75,
            'col_width':  img_w / 7,
        }

    except Exception:
        return {
            'status': 'error',
            'task':   task,
        }


def process_excel(input_bytes, progress_callback=None):
    """
    Main function for Streamlit.
    Downloads ALL images concurrently (CMD-script speed),
    then embeds in chunks to keep RAM flat.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(input_bytes))
    red_font = Font(color="FF0000", bold=True)

    # ── Step 1: Collect all URL tasks (same as your script) ───────────────────
    tasks_to_process = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(('http://', 'https://')):
                    tasks_to_process.append({
                        'sheet_name': sheet.title,
                        'coordinate': cell.coordinate,
                        'url':        cell.value,
                    })

    total = len(tasks_to_process)
    stats = {'success': 0, 'error': 0, 'total': total}

    if total == 0:
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output, stats

    # ── Step 2: Download ALL images concurrently — CMD script speed ───────────
    # Results stored as plain bytes (not openpyxl objects) — RAM-efficient
    all_results = [None] * total
    task_index  = {id(task): i for i, task in enumerate(tasks_to_process)}

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {
            executor.submit(download_and_process_image, task): (i, task)
            for i, task in enumerate(tasks_to_process)
        }
        completed = 0
        for future in concurrent.futures.as_completed(future_map):
            idx, _ = future_map[future]
            all_results[idx] = future.result()
            completed += 1
            # Report progress every 10 images so UI stays responsive
            if progress_callback and (completed % 10 == 0 or completed == total):
                progress_callback(completed, total)

    # ── Step 3: Embed into workbook in chunks — keeps peak RAM flat ───────────
    for chunk_start in range(0, total, EMBED_CHUNK):
        chunk = all_results[chunk_start : chunk_start + EMBED_CHUNK]

        for result in chunk:
            task_info = result['task']
            sheet = workbook[task_info['sheet_name']]
            cell  = sheet[task_info['coordinate']]

            if result['status'] == 'success':
                img_obj        = Image(io.BytesIO(result['img_bytes']))
                img_obj.width  = result['img_w']
                img_obj.height = result['img_h']

                sheet.row_dimensions[cell.row].height             = result['row_height']
                sheet.column_dimensions[cell.column_letter].width = result['col_width']
                cell.value = None
                sheet.add_image(img_obj, cell.coordinate)
                stats['success'] += 1
            else:
                cell.value = "Image is corrupt or failed to download"
                cell.font  = red_font
                stats['error'] += 1

        # Free processed chunk bytes from RAM
        for r in chunk:
            r.pop('img_bytes', None)
        gc.collect()

    # ── Step 4: Save ──────────────────────────────────────────────────────────
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, stats
